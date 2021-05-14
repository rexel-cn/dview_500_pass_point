import argparse




def cover_value(temp_val):
    temp_val = temp_val.replace("ERROR", "ERR")
    temp_val = temp_val.replace("REEOR", "ERR")
    temp_val = temp_val.replace("GO_FORWARD", "GFW")
    temp_val = temp_val.replace("REMOTE", "RM")
    temp_val = temp_val.replace("OPEN", "OP")
    temp_val = temp_val.replace("OPEND", "OPD")
    temp_val = temp_val.replace("SUCCESS", "SUCC")
    temp_val = temp_val.replace("CLOSE", "CL")
    temp_val = temp_val.replace("CLOSED", "CLD")
    temp_val = temp_val.replace(".", "_")
    temp_val = temp_val.replace("HIHIDB", "2HID")
    temp_val = temp_val.replace("HIHILIM", "2HILIM")
    temp_val = temp_val.replace("__", "_")
    temp_val = temp_val.replace("START/STOP", "SCS")
    temp_val = temp_val.replace("START", "STR")
    temp_val = temp_val.replace("STOP", "STP")
    temp_val = temp_val.replace("BYPASS", "BP")
    temp_val = temp_val.replace("RUNTIME", "RT")
    temp_val = temp_val.replace("INTERVAL", "INR")
    temp_val = temp_val.replace("OUT_WATER_TIME", "OWT")
    temp_val = temp_val.replace("IN_WATER_TIME", "IWT")
    temp_val = temp_val.replace("JIANYEXIELOU", "JYXL")
    temp_val = temp_val.replace("RUN", "R")
    temp_val = temp_val.replace("LOLO", "LL")
    temp_val = temp_val.replace("RESET", "RST")
    temp_val = temp_val.replace("HIHI", "HH")
    temp_val = temp_val.replace("LOCAL", "LCA")
    temp_val = temp_val.replace("BLOW_TIME", "BT")
    temp_val = temp_val.replace("MANUAL", "MAN")
    temp_val = temp_val.replace("GAS_OUTLET_BQ", "GOB")
    temp_val = temp_val.replace("BLOW_TIME_BQ", "BOB")
    temp_val = temp_val.replace("RECEIVE", "RECV")
    temp_val = temp_val.replace("(", "")
    temp_val = temp_val.replace(")", "")
    temp_val = temp_val.replace("#", "")
    temp_val = temp_val.replace("_A_", "_")
    return temp_val


if __name__ == '__main__':
    parser = argparse.ArgumentParser(description='Excel transfer')
    parser.add_argument('--config', '-c', help='基础配置文件')
    parser.add_argument('--prod', '-p', help='输出的产品模板')
    parser.add_argument('--device', '-d', help='输出的设备模板')
    parser.add_argument('--point', '-t', help='输出的点位模板')
    args = parser.parse_args()
    from config import sheet_list
    from openpyxl import load_workbook

    file_home = args.config
    out_file_home = str(args.config).replace(".xlsx", "_transfer.xlsx")
    wb = load_workbook(filename=file_home)


    def load_index_cache():
        ws = wb['监控对象']
        montier_obj = {}
        rows = ws.rows
        # 迭代所有的行
        for row in rows:
            line = [col.value for col in row]
            montier_obj[line[0]] = ""
        return montier_obj


    montier_obj = load_index_cache()

    # print(montier_obj)

    for sheet_var in sheet_list:
        column_index = sheet_var.split("=")[1]
        sheet_name = sheet_var.split("=")[0]
        wss = wb[sheet_name]
        # rows = wss.rows
        # columns = wss.columns
        # 迭代所有的行
        for index, row in enumerate(wss.rows, start=2):
            temp_value = wss.cell(row=index, column=int(column_index)).value
            desc_value = wss.cell(row=index, column=int(column_index) + 1).value
            if temp_value == None:
                continue
            # if len(temp_value) >= 20:
            #     print("{} len =  {}  :default max20".format(sheet_name, temp_value))
            #  监控对象和是否属于那种类型
            for key, values in montier_obj.items():
                if key in temp_value:
                    montier_obj[key] = sheet_name

            temp_value = cover_value(temp_value)
            temp_value = "{}_{}".format(sheet_name, temp_value)
            wss.cell(row=index, column=int(column_index)).value = temp_value

        # print(montier_obj)
    ws1 = wb['监控对象']
    not_found_point = 0
    for index1, row1 in enumerate(ws1.rows, start=1):
        temp_val_ = ws1.cell(row=index1, column=int(1)).value
        if temp_val_ != None:
            if temp_val_ in montier_obj:
                if montier_obj.get(temp_val_) != "":
                    ws1.cell(row=index1, column=int(1)).value = cover_value(
                        "{}_{}".format(montier_obj.get(temp_val_), temp_val_))
                else:
                    ws1.cell(row=index1, column=int(1)).value = cover_value("{}".format(temp_val_))
                    print("以下监控对象存在，点位关联信息不存在: {}".format(cover_value("{}".format(temp_val_))))
                    not_found_point = not_found_point + 1
    wb.save(out_file_home)

    print("监控对象在点位表查不到的有[{}]".format(not_found_point))
    print("检查是否超过长度............................")

    print("==============监控对象总数量(不包含重复的)====================")
    print(len(montier_obj.keys()))
    print("===========================================")
    print("开始检查长度超过20的点位名称。")
    wb = load_workbook(filename=out_file_home)
    for sheet_var in sheet_list:  # 8组
        column_index = sheet_var.split("=")[1]
        sheet_name = sheet_var.split("=")[0]
        wss = wb[sheet_name]
        for index, row in enumerate(wss.rows, start=2):
            temp_value = wss.cell(row=index, column=int(column_index)).value
            if temp_value == None:
                continue
            if len(temp_value) >= 20:
                print("{} len =  {}  :default max20".format(sheet_name, temp_value))

    print("开始检测 并且准备生成点位 和 产品数据")

    for key in list(montier_obj.keys()):
        if not montier_obj.get(key):
            del montier_obj[key]
    print(len(montier_obj))
    print("==================================================")
    # 948个对象
    gl_all_data_kv = {}
    for cover_sheet_var in sheet_list:
        print("开始处理 = " + cover_sheet_var)
        column_index = cover_sheet_var.split("=")[1]
        sheet_name = cover_sheet_var.split("=")[0]
        wss = wb[sheet_name]
        for index, row in enumerate(wss.rows, start=2):
            temp_value = wss.cell(row=index, column=int(column_index)).value
            desc_value = wss.cell(row=index, column=int(column_index) + 1).value
            if temp_value != None:
                for key, values in montier_obj.items():
                    # 例如 key  =  K1_GP1
                    # 例如 value  = DR
                    if key in temp_value:
                        mk = montier_obj[key]
                        tm_key = "{}_{}".format(mk, key)
                        if gl_all_data_kv.get(tm_key):
                            temp_a = gl_all_data_kv.get(tm_key)
                            temp_a.append("{}###{}".format(tm_key, desc_value))
                            gl_all_data_kv[tm_key] = temp_a
                        else:
                            gl_all_data_kv[tm_key] = ["{}###{}".format(tm_key, desc_value)]

    print("gl_all_data_kv len = " + str(len(gl_all_data_kv)))
    print("生成产品表")
    print(len(montier_obj))
    wb1 = load_workbook(filename="data\\prod.xlsx")
    wss = wb1["产品信息"]
    index = 2

    for key in gl_all_data_kv.keys():
        wss.cell(row=index, column=int(2)).value = key
        wss.cell(row=index, column=int(4)).value = str(gl_all_data_kv.get(key)[0]).split("###")[1]
        index = index + 1

    wb1.save(filename="data\\prod_out.xlsx")

    wb1 = load_workbook(filename="data\\point.xlsx")
    wss = wb1["point_tag"]
    index = 2
    for key in gl_all_data_kv.keys():
        for m_arry in gl_all_data_kv.get(key):
            wss.cell(row=index, column=int(1)).value = key
            wss.cell(row=index, column=int(2)).value = str(gl_all_data_kv.get(key)[0]).split("###")[1]
            wss.cell(row=index, column=int(4)).value = str(m_arry).split("###")[0]
            wss.cell(row=index, column=int(5)).value = str(m_arry).split("###")[1]
            index = index + 1

    wb1.save(filename="data\\point_out.xlsx")
